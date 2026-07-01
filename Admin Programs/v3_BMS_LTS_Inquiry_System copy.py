import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from datetime import datetime
from tkcalendar import DateEntry
import json
import sys
import re

# Path to JSON config containing process_flow and process_column_mapping
config_file_path = r"\\phlsvr08\BMS Data\BMS_Database\Lot_Tracking\process_flow.json"
# config_file_path = "/path/to/process_flow.json"  # alternative path for linux/mac

def show_error_and_exit(title, msg):
    tmp = tk.Tk()
    tmp.withdraw()
    messagebox.showerror(title, msg)
    tmp.destroy()
    sys.exit(1)

# Load process_flow and process_column_mapping from JSON
try:
    with open(config_file_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    process_flow = config.get("process_flow")
    process_column_mapping = config.get("process_column_mapping")

    if not isinstance(process_flow, list) or not isinstance(process_column_mapping, dict):
        raise ValueError("JSON must contain 'process_flow' (list) and 'process_column_mapping' (dict).")

    # Validate mapping values contain expected indices (we use indices 0..4 in this script)
    for proc, cols in process_column_mapping.items():
        if not isinstance(cols, (list, tuple)) or len(cols) < 5:
            raise ValueError(f"Mapping for process '{proc}' must be a list/tuple with at least 5 elements "
                             "(input, output, defect, remarks, proc_date).")
except FileNotFoundError:
    show_error_and_exit("Configuration Error", f"Config file not found at: {config_file_path}")
except json.JSONDecodeError:
    show_error_and_exit("Configuration Error", f"Error decoding JSON config at: {config_file_path}")
except Exception as e:
    show_error_and_exit("Configuration Error", f"Invalid configuration: {e}")

# Database file
DB_FILE = r"\\phlsvr08\BMS Data\BMS_Database\Lot_Tracking\lot_tracking.db"


def query_lot_details(lot_number=None, sensor_id=None, date_from=None, date_to=None):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Build the initial query
    query = "SELECT lot_number, sensor_id, lot_entry_proc_date FROM lot_tracking WHERE 1=1"
    params = []

    # Add conditions for lot_number and sensor_id if provided
    if lot_number:
        query += " AND lot_number = ?"
        params.append(lot_number)
    if sensor_id:
        query += " AND sensor_id = ?"
        params.append(sensor_id)
    
    # Date range search logic, if date_from and date_to are provided
    if date_from and date_to:
        start_date = date_from + " 00:00:00"
        end_date = date_to + " 23:59:59"
        query += " AND lot_entry_proc_date BETWEEN ? AND ?"
        params.extend([start_date, end_date])

    cursor.execute(query, params)
    result = cursor.fetchall()
    conn.close()
    return result


# Utility: basic column name validation (defensive)
def is_valid_column_name(name):
    return bool(re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', str(name)))


def delete_defects_remarks_for_sensor(sensor_id, lot_number=None, clear_dates=True):
    """
    Clear defect and remark columns (and optionally proc_date) for a given sensor_id.
    If lot_number is provided, updates only rows matching that lot_number.
    Returns number of rows changed (total rows updated).
    """
    if not sensor_id:
        return 0

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    # Track total changes in this connection
    initial_total = conn.total_changes

    for process, cols in process_column_mapping.items():
        # Defensive: need at least indices 2 (defect), 3 (remark)
        if not cols or len(cols) < 4:
            continue

        defect_col = cols[2]
        remark_col = cols[3]
        proc_date_col = cols[4] if len(cols) > 4 else None

        # Validate column names
        if not (is_valid_column_name(defect_col) and is_valid_column_name(remark_col)):
            continue
        if proc_date_col and not is_valid_column_name(proc_date_col):
            proc_date_col = None

        # Build the UPDATE statement
        set_clauses = [f"{defect_col} = NULL", f"{remark_col} = NULL"]
        if clear_dates and proc_date_col:
            set_clauses.append(f"{proc_date_col} = NULL")
        set_clause = ", ".join(set_clauses)

        query = f"UPDATE lot_tracking SET {set_clause} WHERE sensor_id = ?"
        params = [sensor_id]
        if lot_number:
            query += " AND lot_number = ?"
            params.append(lot_number)

        try:
            cursor.execute(query, params)
        except sqlite3.OperationalError:
            # Column may not exist in DB; skip this process
            continue

    conn.commit()
    total_changed = conn.total_changes - initial_total
    conn.close()
    return total_changed


# Function to fetch details based on Lot Number, Sensor ID, or Date Range
def fetch_details():
    search_type = search_type_combobox.get()

    if search_type == "Date":
        search_from_date = from_date_entry.get()
        search_to_date = to_date_entry.get()
        
        if not search_from_date or not search_to_date:
            messagebox.showwarning("Input Error", "Please enter both 'From' and 'To' dates in MM/DD/YYYY format.")
            return

        try:
            # Convert the input dates to datetime objects
            search_from_date_obj = datetime.strptime(search_from_date, "%m/%d/%Y")
            search_to_date_obj = datetime.strptime(search_to_date, "%m/%d/%Y")
        except ValueError:
            messagebox.showwarning("Invalid Date Format", "Please enter the dates in the format MM/DD/YYYY.")
            return
        
        # Format the datetime objects to strings for the date part
        search_from_date_str = search_from_date_obj.strftime("%m/%d/%Y")
        search_to_date_str = search_to_date_obj.strftime("%m/%d/%Y")
        
        # Use the query_lot_details function to search the database
        records = query_lot_details(date_from=search_from_date_str, date_to=search_to_date_str)
        
        if not records:
            messagebox.showinfo("No Results", f"No records found between {search_from_date_str} and {search_to_date_str}.")
            return
        
        # Clear the sensor list before adding new search results
        sensor_list.delete(*sensor_list.get_children())
        last_sensor_data.clear()
        
        for row in records:
            lot_number, sensor_id, lot_entry_proc_date = row
            
            # Now, we will search for the defects and remarks for this sensor_id for all processes
            defects = []
            remarks = []
            defect_processes = []
            remark_processes = []
            defect_dates = []
            remark_dates = []
            
            # Iterate over the processes to get defect and remark details
            for process, columns in process_column_mapping.items():
                # Defensive: ensure columns list has required indices
                if not columns or len(columns) < 5:
                    continue

                conn = sqlite3.connect(DB_FILE)
                cursor = conn.cursor()
                try:
                    cursor.execute(f"""
                        SELECT {columns[2]}, {columns[3]}, {columns[4]} 
                        FROM lot_tracking 
                        WHERE lot_number = ? AND sensor_id = ? AND lot_entry_proc_date = ?
                    """, (lot_number, sensor_id, lot_entry_proc_date))
                    result = cursor.fetchone()
                except sqlite3.OperationalError:
                    # If the database doesn't have these columns, skip
                    result = None
                finally:
                    conn.close()
                
                if result:
                    defect, remark, proc_date = result
                    if defect:
                        defects.append(defect)
                        defect_processes.append(process)
                        defect_dates.append(proc_date if proc_date else "")
                    if remark:
                        remarks.append(remark)
                        remark_processes.append(process)
                        remark_dates.append(proc_date if proc_date else "")
            
            # Prepare the defects and remarks for display (leave empty if no data)
            defects_str = "; ".join(defects) if defects else ""
            remarks_str = "; ".join(remarks) if remarks else ""
            defect_processes_str = "; ".join(defect_processes) if defect_processes else ""
            remark_processes_str = "; ".join(remark_processes) if remark_processes else ""
            defect_dates_str = "; ".join(defect_dates) if defect_dates else ""
            remark_dates_str = "; ".join(remark_dates) if remark_dates else ""
            
            # Insert the data into the sensor list table and cache it
            vals = (lot_number, sensor_id, defects_str, defect_processes_str,
                    defect_dates_str, remarks_str, remark_processes_str, remark_dates_str)
            if defects_str or remarks_str:
                sensor_list.insert("", "end", values=vals, tags=('excluded',))
            else:
                sensor_list.insert("", "end", values=vals)
            last_sensor_data.append(vals)
        
        return

    search_value = search_value_entry.get().strip()
    if not search_value:
        messagebox.showwarning("Input Error", "Please enter a value to search.")
        return
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    if search_type == "Lot Number":
        cursor.execute("SELECT lot_number, current_process, lot_entry_IN FROM lot_tracking WHERE lot_number=? LIMIT 1", (search_value,))
    elif search_type == "Sensor ID":
        cursor.execute("SELECT lot_number, current_process, lot_entry_IN FROM lot_tracking WHERE sensor_id=? LIMIT 1", (search_value,))
    
    row = cursor.fetchone()
    
    if not row:
        messagebox.showinfo("No Results", "No records found.")
        conn.close()
        return
    
    details_list.delete(*details_list.get_children())
    
    lot_number, current_process, lot_entry_IN = row

    # Determine previous process output defensively
    previous_output = "N/A"
    try:
        current_process_index = process_flow.index(current_process)
    except ValueError:
        current_process_index = -1

    if current_process_index > 0:
        previous_process = process_flow[current_process_index - 1]
        previous_columns = process_column_mapping.get(previous_process)
        if previous_columns and len(previous_columns) >= 2:
            try:
                cursor.execute(f"SELECT {previous_columns[1]} FROM lot_tracking WHERE lot_number=? LIMIT 1", (lot_number,))
                prev_row = cursor.fetchone()
                previous_output = prev_row[0] if prev_row and prev_row[0] is not None else "N/A"
            except sqlite3.OperationalError:
                previous_output = "N/A"
        else:
            previous_output = "N/A"
    else:
        previous_output = "N/A"
    
    details_list.insert("", "end", values=(lot_number, current_process, lot_entry_IN, previous_output))
    
    # Fetch sensor details
    sensor_list.delete(*sensor_list.get_children())
    last_sensor_data.clear()
    
    if search_type == "Sensor ID":
        sensor_ids = [(search_value,)]
    else:
        cursor.execute("SELECT sensor_id FROM lot_tracking WHERE lot_number=?", (lot_number,))
        sensor_ids = cursor.fetchall()
    
    for sensor_id_tuple in sensor_ids:
        sensor_id = sensor_id_tuple[0]
        defects = []
        remarks = []
        defect_processes = []
        remark_processes = []
        defect_dates = []
        remark_dates = []
        
        for process, columns in process_column_mapping.items():
            if not columns or len(columns) < 5:
                continue
            try:
                cursor.execute(f"SELECT {columns[2]}, {columns[3]}, {columns[4]} FROM lot_tracking WHERE lot_number=? AND sensor_id=?", (lot_number, sensor_id))
                result = cursor.fetchone()
            except sqlite3.OperationalError:
                result = None
            if result:
                defect, remark, proc_date = result
                if defect:
                    defects.append(defect)
                    defect_processes.append(process)
                    defect_dates.append(proc_date if proc_date else "")
                if remark:
                    remarks.append(remark)
                    remark_processes.append(process)
                    remark_dates.append(proc_date if proc_date else "")
        
        defects_str = "; ".join(defects)
        remarks_str = "; ".join(remarks)
        defect_processes_str = "; ".join(defect_processes)
        remark_processes_str = "; ".join(remark_processes)
        defect_dates_str = "; ".join(defect_dates)
        remark_dates_str = "; ".join(remark_dates)
        
        vals = (lot_number, sensor_id, defects_str, defect_processes_str, defect_dates_str, remarks_str, remark_processes_str, remark_dates_str)
        if defects_str or remarks_str:
            sensor_list.insert("", "end", values=vals, tags=('excluded',))
        else:
            sensor_list.insert("", "end", values=vals)
        last_sensor_data.append(vals)
    
    conn.close()
    
    # Auto-fit column widths
    details_list.update_idletasks()
    sensor_list.update_idletasks()


# Function to clear all fields
def clear_fields():
    search_value_entry.delete(0, tk.END)
    details_list.delete(*details_list.get_children())
    sensor_list.delete(*sensor_list.get_children())
    last_sensor_data.clear()


# Function to export data to an Excel file
def export_to_excel():
    # Get data from details_list
    details_data = [details_list.item(item)["values"] for item in details_list.get_children()]
    # Get data from sensor_list
    sensor_data = [sensor_list.item(item)["values"] for item in sensor_list.get_children()]
    
    # Check if there is data to export
    if not details_data and not sensor_data:
        messagebox.showwarning("No Data", "No data to export.")
        return
    
    # Prompt user to select a location and file name to save the file
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Save as")
    
    if not file_path:
        return  # If the user cancels the dialog, do nothing
    
    # Create a pandas Excel writer
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        # Write details_list data to 'Details' sheet
        if details_data:
            details_df = pd.DataFrame(details_data, columns=["Lot Number", "Current Process", "Input", "Output"])
            details_df.to_excel(writer, sheet_name="Details", index=False)
        
        # Write sensor_list data to 'Sensor Details' sheet
        if sensor_data:
            sensor_df = pd.DataFrame(sensor_data, columns=sensor_columns)
            sensor_df.to_excel(writer, sheet_name="Sensor Details", index=False)
    
    # Load the workbook to apply formatting
    workbook = load_workbook(file_path)
    
    # Remove borders on headers and autofit columns
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Remove borders from the header row
        for cell in sheet[1]:
            cell.border = Border(left=Side(border_style=None),
                                 right=Side(border_style=None),
                                 top=Side(border_style=None),
                                 bottom=Side(border_style=None))

        # Autofit columns based on the content
        for column in sheet.columns:
            max_length = 0
            col = column[0].column_letter  # Get the column letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # Adjust width (1.2 added to give some padding)
            sheet.column_dimensions[col].width = max_length + 1.2
    
    # Save the workbook with the new formatting
    workbook.save(file_path)
    
    messagebox.showinfo("Export Successful", f"Data has been exported to {file_path}")


def delete_selected_sensor():
    # Get selected row from sensor_list
    sel = sensor_list.selection()
    if not sel:
        messagebox.showwarning("Selection Required", "Please select a sensor row from the list to delete defects/remarks.")
        return

    item = sensor_list.item(sel[0])
    values = item.get("values", [])
    if len(values) < 2:
        messagebox.showwarning("Selection Error", "Selected row does not contain expected values.")
        return

    lot_number = values[0]
    sensor_id = values[1]

    # Confirm action
    if not messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete defects/remarks for Sensor ID '{sensor_id}' (Lot '{lot_number}')? This cannot be undone."):
        return

    # Optionally ask whether to also clear proc_date
    clear_dates = messagebox.askyesno("Clear Dates?", "Also clear the process date fields associated with defects/remarks?")

    try:
        updated = delete_defects_remarks_for_sensor(sensor_id, lot_number=lot_number, clear_dates=clear_dates)
        messagebox.showinfo("Operation Complete", f"Update complete. {updated} row(s) affected.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while updating the database:\n{e}")
        return

    # Refresh the current search results so the UI shows updated values.
    fetch_details()


# Setting up the GUI
root = tk.Tk()
root.title("BMS Lot Tracking Query System")
root.geometry("1210x550")
root.configure(bg='#3a6ba8')
# Disable maximize functionality
root.resizable(False, False)

# Labels and Entries
tk.Label(root, text="Search By:", bg='#3a6ba8', fg="white").place(x=5, y=15)
search_type_combobox = ttk.Combobox(root, values=["Lot Number", "Sensor ID", "Date"], width=27)
search_type_combobox.place(x=105, y=15)
search_type_combobox.set("Lot Number")
search_type_combobox.bind("<<ComboboxSelected>>", lambda event: clear_fields())

tk.Label(root, text="Value:", bg='#3a6ba8', fg="white").place(x=5, y=45)
search_value_entry = tk.Entry(root, width=30)
search_value_entry.place(x=105, y=45)

# Search Button
search_button = tk.Button(root, text="Search", command=fetch_details, bg="green", fg="white", font=("Tahoma", 12, "bold"), padx=10, pady=2, relief='raised', borderwidth=3)
search_button.place(x=550, y=10)

# Clear Button
clear_button = tk.Button(root, text="Clear", command=clear_fields, bg="red", fg="white", font=("Tahoma", 12, "bold"), padx=17, pady=1, relief='raised', borderwidth=3)
clear_button.place(x=550, y=50)

# Export Button
export_button = tk.Button(root, text="Export", command=export_to_excel, bg="blue", fg="white", font=("Tahoma", 12, "bold"), padx=12, pady=1, relief='raised', borderwidth=3)
export_button.place(x=550, y=90)

# Show excluded sensors (have defects or remarks)
def show_excluded():
    global last_sensor_data
    if not last_sensor_data:
        messagebox.showinfo("No Data", "No sensor data available to filter. Run a search first.")
        return
    sensor_list.delete(*sensor_list.get_children())
    for vals in last_sensor_data:
        # vals order matches sensor_columns
        defects = vals[2]
        remarks = vals[5]
        if defects or remarks:
            sensor_list.insert("", "end", values=vals, tags=('excluded',))

# Restore all sensor rows from cache
def show_all():
    global last_sensor_data
    if not last_sensor_data:
        messagebox.showinfo("No Data", "No sensor data available. Run a search first.")
        return
    sensor_list.delete(*sensor_list.get_children())
    for vals in last_sensor_data:
        if vals[2] or vals[5]:
            sensor_list.insert("", "end", values=vals, tags=('excluded',))
        else:
            sensor_list.insert("", "end", values=vals)

# Show Excluded / All buttons
show_excluded_btn = tk.Button(root, text="Show Excluded", command=show_excluded, bg="#aa0000", fg="white", font=("Tahoma", 10, "bold"), padx=8, pady=1, relief='raised', borderwidth=2)
show_excluded_btn.place(x=830, y=10)
show_all_btn = tk.Button(root, text="Show All", command=show_all, bg="#0077aa", fg="white", font=("Tahoma", 10, "bold"), padx=8, pady=1, relief='raised', borderwidth=2)
show_all_btn.place(x=830, y=50)

# Delete Button (new)
delete_button = tk.Button(root, text="Delete Defects/Remarks", command=delete_selected_sensor, bg="orange", fg="white", font=("Tahoma", 10, "bold"), padx=10, pady=2, relief='raised', borderwidth=3)
delete_button.place(x=700, y=10)

# Search by Date UI elements
tk.Label(root, text="From Date:", bg='#3a6ba8', fg="white").place(x=350, y=15)
from_date_entry = DateEntry(root, width=12, date_pattern="mm/dd/yyyy")
from_date_entry.place(x=420, y=15)

tk.Label(root, text="To Date:", bg='#3a6ba8', fg="white").place(x=350, y=45)
to_date_entry = DateEntry(root, width=12, date_pattern="mm/dd/yyyy")
to_date_entry.place(x=420, y=45)

# Table for displaying details
columns = ("Lot Number", "Current Process", "Input", "Previous Output")
details_list = ttk.Treeview(root, columns=columns, show="headings", height=10)
details_list.heading("Lot Number", text="Lot Number")
details_list.heading("Current Process", text="Current Process")
details_list.heading("Input", text="Input")
details_list.heading("Previous Output", text="Output")
details_list.column("Lot Number", width=50, stretch=True)
details_list.column("Current Process", width=150, stretch=True)
details_list.column("Input", width=5, stretch=True)
details_list.column("Previous Output", width=5, stretch=True)
details_list.place(x=5, y=80, width=500, height=50)

# Define the modified sensor columns
sensor_columns = ("Lot Number", "Sensor ID", "Defects", "Process of Defect", "Date of Defect", "Remarks", "Process of Remarks", "Date of Remarks")
sensor_list = ttk.Treeview(root, columns=sensor_columns, show="headings", height=10)
sensor_list.heading("Lot Number", text="Lot Number")
sensor_list.heading("Sensor ID", text="Sensor ID")
sensor_list.heading("Defects", text="Defects")
sensor_list.heading("Remarks", text="Remarks")
sensor_list.heading("Process of Defect", text="Process of Defect")
sensor_list.heading("Process of Remarks", text="Process of Remarks")
sensor_list.heading("Date of Defect", text="Date of Defect")
sensor_list.heading("Date of Remarks", text="Date of Remarks")
sensor_list.column("Lot Number", width=90, stretch=True)
sensor_list.column("Sensor ID", width=90, stretch=True)
sensor_list.column("Defects", width=100, stretch=True)
sensor_list.column("Remarks", width=100, stretch=True)
sensor_list.column("Process of Defect", width=150, stretch=True)
sensor_list.column("Process of Remarks", width=150, stretch=True)
sensor_list.column("Date of Defect", width=100, stretch=True)
sensor_list.column("Date of Remarks", width=100, stretch=True)
sensor_list.place(x=5, y=140, width=1200, height=400)

# Cache last populated sensor rows so we can filter/extract excluded sensors quickly
last_sensor_data = []  # list of tuples matching sensor_columns order

# Tag for excluded sensors (have defect or remark)
sensor_list.tag_configure('excluded', background='#ffdddd')

# Start the GUI event loop
root.mainloop()