"""
Material Request Sheet System
Admin program for creating and managing material request sheets for production
Based on customer PO and production requirements
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.drawing.image import Image as XLImage
import os
import json
import barcode
from io import BytesIO
from PIL import Image, ImageTk

# Database paths
DB_BASE_PATH = r"\\phlsvr08\BMS Data\BMS_Database\Lot_Tracking"
REQUEST_SHEET_DB = os.path.join(DB_BASE_PATH, "material_request_sheets.db")

class MaterialRequestSheetSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Material Request Sheet System")
        self.root.geometry("1400x800")
        self.root.configure(bg='lightgray')
        self.root.resizable(True, True)
        
        # Initialize database
        self.init_database()
        
        # Track material rows
        self.material_rows = []
        self.next_row_number = 1
        
        # Build UI
        self.create_ui()
    
    def init_database(self):
        """Initialize SQLite database for request sheets"""
        conn = sqlite3.connect(REQUEST_SHEET_DB)
        cursor = conn.cursor()
        
        # Request sheets master table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS request_sheets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_sheet_no TEXT UNIQUE NOT NULL,
                issue_date TEXT NOT NULL,
                customer_name TEXT NOT NULL,
                qty_of_sensor INTEGER NOT NULL,
                po_number TEXT NOT NULL,
                item_code TEXT,
                checker_name TEXT,
                checker_date TEXT,
                approver_name TEXT,
                approver_date TEXT,
                created_by TEXT NOT NULL,
                created_date TEXT NOT NULL,
                status TEXT DEFAULT 'Pending'
            )
        """)
        
        # Material items detail table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS material_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_sheet_no TEXT NOT NULL,
                item_no INTEGER NOT NULL,
                model_no TEXT NOT NULL,
                barcode TEXT,
                qty_weight TEXT NOT NULL,
                FOREIGN KEY (request_sheet_no) REFERENCES request_sheets(request_sheet_no)
            )
        """)
        
        conn.commit()
        conn.close()
    
    def create_ui(self):
        """Create the main user interface"""
        # Title and toolbar
        title_frame = tk.Frame(self.root, bg='lightgray')
        title_frame.pack(fill='x', padx=15, pady=15)
        
        tk.Label(title_frame, text="Material Request Sheet System", 
                font=("Segoe UI Semibold", 24, "bold"), bg="lightgray", fg="#1e293b").pack(side='left')
        
        # Toolbar buttons
        toolbar = tk.Frame(title_frame, bg='lightgray')
        toolbar.pack(side='right')
        
        tk.Button(toolbar, text="📋 View All Sheets", command=self.show_request_list_popup, bg="#6366f1", fg="white",
                 font=("Segoe UI", 11, "bold"), padx=15, pady=5, relief='raised', borderwidth=2).pack(side='left', padx=5)
        
        tk.Button(toolbar, text="➕ New Sheet", command=self.new_request_sheet, bg="#16a34a", fg="white",
                 font=("Segoe UI", 11, "bold"), padx=15, pady=5, relief='raised', borderwidth=2).pack(side='left', padx=5)
        
        # Main container - Request sheet form only
        main_container = tk.Frame(self.root, bg='lightgray')
        main_container.pack(fill='both', expand=True, padx=15, pady=(0, 15))
        
        # Request sheet form panel
        form_panel = tk.LabelFrame(main_container, text="Request Sheet Details", bg='white',
                                    font=("Segoe UI Semibold", 12), padx=20, pady=20)
        form_panel.pack(fill='both', expand=True)
        
        # Header section
        header_frame = tk.Frame(form_panel, bg='white')
        header_frame.pack(fill='x', pady=(0, 15))
        
        # Left side - Customer Info
        left_info = tk.Frame(header_frame, bg='white')
        left_info.pack(side='left', fill='both', expand=True)
        
        tk.Label(left_info, text="Customer Information", bg='white', 
                font=("Segoe UI Semibold", 12), fg="#1e293b").grid(row=0, column=0, columnspan=2, sticky='w', pady=(0, 10))
        
        tk.Label(left_info, text="Customer Name:", bg='white', font=("Segoe UI", 10)).grid(row=1, column=0, sticky='w', pady=8)
        self.customer_name_entry = tk.Entry(left_info, width=35, font=("Segoe UI", 10))
        self.customer_name_entry.grid(row=1, column=1, sticky='w', padx=(10, 0), pady=8)
        
        tk.Label(left_info, text="Qty of Sensor:", bg='white', font=("Segoe UI", 10)).grid(row=2, column=0, sticky='w', pady=8)
        self.qty_sensor_entry = tk.Entry(left_info, width=35, font=("Segoe UI", 10))
        self.qty_sensor_entry.grid(row=2, column=1, sticky='w', padx=(10, 0), pady=8)
        
        tk.Label(left_info, text="PO Number:", bg='white', font=("Segoe UI", 10)).grid(row=3, column=0, sticky='w', pady=8)
        self.po_number_entry = tk.Entry(left_info, width=35, font=("Segoe UI", 10))
        self.po_number_entry.grid(row=3, column=1, sticky='w', padx=(10, 0), pady=8)
        
        tk.Label(left_info, text="Item Code:", bg='white', font=("Segoe UI", 10)).grid(row=4, column=0, sticky='w', pady=8)
        self.item_code_entry = tk.Entry(left_info, width=35, font=("Segoe UI", 10))
        self.item_code_entry.grid(row=4, column=1, sticky='w', padx=(10, 0), pady=8)
        
        # Right side - Request Info
        right_info = tk.Frame(header_frame, bg='white')
        right_info.pack(side='right', fill='both', padx=(30, 0))
        
        tk.Label(right_info, text="Request Sheet Info", bg='white',
                font=("Segoe UI Semibold", 12), fg="#1e293b").grid(row=0, column=0, columnspan=2, sticky='e', pady=(0, 10))
        
        tk.Label(right_info, text="Issue Date:", bg='white', font=("Segoe UI", 10)).grid(row=1, column=0, sticky='e', pady=8, padx=(0, 10))
        self.issue_date_entry = tk.Entry(right_info, width=25, font=("Segoe UI", 10))
        self.issue_date_entry.grid(row=1, column=1, sticky='e', pady=8)
        self.issue_date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        tk.Label(right_info, text="Request Sheet No:", bg='white', font=("Segoe UI", 10)).grid(row=2, column=0, sticky='e', pady=8, padx=(0, 10))
        self.request_no_entry = tk.Entry(right_info, width=25, font=("Segoe UI", 10))
        self.request_no_entry.grid(row=2, column=1, sticky='e', pady=8)
        self.generate_request_number()
        
        # Separator
        ttk.Separator(form_panel, orient='horizontal').pack(fill='x', pady=15)
        
        # Material List section
        material_header = tk.Frame(form_panel, bg='white')
        material_header.pack(fill='x', pady=(0, 10))
        
        tk.Label(material_header, text="Material List", bg='white',
                font=("Segoe UI Semibold", 12), fg="#1e293b").pack(side='left')
        
        tk.Button(material_header, text="+ Add Material", command=self.add_material_row, 
                 bg="#2563eb", fg="white", font=("Segoe UI", 10, "bold"), 
                 padx=15, pady=5, relief='raised', borderwidth=2).pack(side='right')
        
        # Material table frame with scrollbar
        table_container = tk.Frame(form_panel, bg='white')
        table_container.pack(fill='both', expand=True, pady=(0, 15))
        
        # Canvas for scrollable material list
        canvas = tk.Canvas(table_container, bg='white', highlightthickness=0)
        scrollbar_y = ttk.Scrollbar(table_container, orient="vertical", command=canvas.yview)
        self.material_frame = tk.Frame(canvas, bg='white')
        
        self.material_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.material_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_y.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        
        # Material table headers
        headers = ["No", "Model No", "Barcode (Model No)", "Qty/Weight", "Action"]
        widths = [5, 25, 30, 15, 10]
        
        for idx, (header, width) in enumerate(zip(headers, widths)):
            tk.Label(self.material_frame, text=header, bg='#2563eb', fg='white',
                    font=("Segoe UI Semibold", 10), relief='solid', borderwidth=1,
                    width=width).grid(row=0, column=idx, sticky='ew', padx=1, pady=1)
        
        # Separator
        ttk.Separator(form_panel, orient='horizontal').pack(fill='x', pady=15)
        
        # Approval section - Better layout
        approval_label = tk.Label(form_panel, text="Approval Section", bg='white',
                                 font=("Segoe UI Semibold", 12), fg="#1e293b")
        approval_label.pack(anchor='w', pady=(0, 10))
        
        approval_frame = tk.Frame(form_panel, bg='white')
        approval_frame.pack(fill='x', pady=(0, 15))
        
        # Checker section
        checker_frame = tk.LabelFrame(approval_frame, text="Checked By", bg='white', 
                                      font=("Segoe UI Semibold", 10), padx=15, pady=10, relief='ridge', borderwidth=2)
        checker_frame.pack(side='left', padx=(0, 20), fill='both', expand=True)
        
        tk.Label(checker_frame, text="Name:", bg='white', font=("Segoe UI", 10)).grid(row=0, column=0, sticky='w', pady=5, padx=(0, 10))
        self.checker_name_entry = tk.Entry(checker_frame, width=25, font=("Segoe UI", 10))
        self.checker_name_entry.grid(row=0, column=1, sticky='ew', pady=5)
        
        tk.Label(checker_frame, text="Date:", bg='white', font=("Segoe UI", 10)).grid(row=1, column=0, sticky='w', pady=5, padx=(0, 10))
        self.checker_date_entry = tk.Entry(checker_frame, width=25, font=("Segoe UI", 10))
        self.checker_date_entry.grid(row=1, column=1, sticky='ew', pady=5)
        
        checker_frame.grid_columnconfigure(1, weight=1)
        
        # Approver section
        approver_frame = tk.LabelFrame(approval_frame, text="Approved By", bg='white',
                                       font=("Segoe UI Semibold", 10), padx=15, pady=10, relief='ridge', borderwidth=2)
        approver_frame.pack(side='left', fill='both', expand=True)
        
        tk.Label(approver_frame, text="Name:", bg='white', font=("Segoe UI", 10)).grid(row=0, column=0, sticky='w', pady=5, padx=(0, 10))
        self.approver_name_entry = tk.Entry(approver_frame, width=25, font=("Segoe UI", 10))
        self.approver_name_entry.grid(row=0, column=1, sticky='ew', pady=5)
        
        tk.Label(approver_frame, text="Date:", bg='white', font=("Segoe UI", 10)).grid(row=1, column=0, sticky='w', pady=5, padx=(0, 10))
        self.approver_date_entry = tk.Entry(approver_frame, width=25, font=("Segoe UI", 10))
        self.approver_date_entry.grid(row=1, column=1, sticky='ew', pady=5)
        
        approver_frame.grid_columnconfigure(1, weight=1)
        
        # Button panel
        button_frame = tk.Frame(form_panel, bg='white')
        button_frame.pack(fill='x', pady=(10, 0))
        
        tk.Button(button_frame, text="💾 Save", command=self.save_request_sheet, bg="#16a34a", fg="white",
                 font=("Segoe UI", 12, "bold"), padx=25, pady=8, relief='raised', borderwidth=3).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="📊 Export to Excel", command=self.export_to_excel, bg="#0ea5e9", fg="white",
                 font=("Segoe UI", 12, "bold"), padx=25, pady=8, relief='raised', borderwidth=3).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="🗑️ Delete", command=self.delete_request_sheet, bg="#dc2626", fg="white",
                 font=("Segoe UI", 12, "bold"), padx=25, pady=8, relief='raised', borderwidth=3).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="🔄 Clear", command=self.clear_form, bg="#f59e0b", fg="white",
                 font=("Segoe UI", 12, "bold"), padx=25, pady=8, relief='raised', borderwidth=3).pack(side='left', padx=5)
    
    def generate_request_number(self):
        """Generate unique request sheet number"""
        date_prefix = datetime.now().strftime("%Y%m%d")
        
        conn = sqlite3.connect(REQUEST_SHEET_DB)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM request_sheets WHERE request_sheet_no LIKE ?", (f"RS-{date_prefix}%",))
        count = cursor.fetchone()[0]
        conn.close()
        
        request_no = f"RS-{date_prefix}-{count + 1:03d}"
        self.request_no_entry.delete(0, tk.END)
        self.request_no_entry.insert(0, request_no)
        return request_no
    
    def generate_barcode(self, model_no):
        """Generate barcode image from model number"""
        try:
            # Import ImageWriter only when needed to avoid compatibility issues
            from barcode.writer import ImageWriter
            
            # Create barcode using the same method as in the project
            code128 = barcode.get_barcode_class('code128')
            writer = ImageWriter()
            barcode_instance = code128(str(model_no), writer=writer)
            
            buffer = BytesIO()
            barcode_instance.write(buffer)
            buffer.seek(0)
            
            # Load image
            img = Image.open(buffer)
            if img.mode != 'RGB':
                img = img.convert('RGB')
            # Resize to fit in the entry field area
            img = img.resize((200, 60), Image.Resampling.LANCZOS)
            
            return img
        except ImportError as ie:
            print(f"Barcode library import error: {ie}")
            print("Barcode generation disabled - please update python-barcode library")
            return None
        except Exception as e:
            print(f"Barcode generation error: {e}")
            return None
    
    def on_model_no_change(self, event, row_data):
        """Auto-generate barcode when model number changes"""
        model_no = event.widget.get().strip()
        if model_no:
            # Update barcode entry with the model number
            row_data['widgets']['barcode'].delete(0, tk.END)
            row_data['widgets']['barcode'].insert(0, model_no)
    
    def show_request_list_popup(self):
        """Show request sheets list in a popup window"""
        popup = tk.Toplevel(self.root)
        popup.title("Request Sheets List")
        popup.geometry("900x600")
        popup.configure(bg='white')
        popup.transient(self.root)
        popup.grab_set()
        
        # Title
        title_frame = tk.Frame(popup, bg='white')
        title_frame.pack(fill='x', padx=15, pady=15)
        
        tk.Label(title_frame, text="Request Sheets", 
                font=("Segoe UI Semibold", 18, "bold"), bg="white", fg="#1e293b").pack(side='left')
        
        # Search frame
        search_frame = tk.Frame(popup, bg='white')
        search_frame.pack(fill='x', padx=15, pady=(0, 10))
        
        tk.Label(search_frame, text="Search:", bg='white', font=("Segoe UI", 10)).pack(side='left', padx=(0, 5))
        search_entry = tk.Entry(search_frame, width=30, font=("Segoe UI", 10))
        search_entry.pack(side='left', padx=(0, 5))
        
        def search_requests():
            search_term = search_entry.get().strip()
            refresh_list(search_term)
        
        search_entry.bind('<KeyRelease>', lambda e: search_requests())
        
        tk.Button(search_frame, text="🔍 Search", command=search_requests, bg="#2563eb", fg="white",
                 font=("Segoe UI", 10, "bold"), padx=15, relief='raised', borderwidth=2).pack(side='left', padx=5)
        
        # Request sheets list
        list_frame = tk.Frame(popup, bg='white')
        list_frame.pack(fill='both', expand=True, padx=15, pady=(0, 15))
        
        columns = ("Request No", "Issue Date", "Customer", "PO Number", "Status")
        request_tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=20)
        
        column_widths = {"Request No": 180, "Issue Date": 120, "Customer": 200, "PO Number": 150, "Status": 100}
        for col in columns:
            request_tree.heading(col, text=col)
            request_tree.column(col, width=column_widths.get(col, 120))
        
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=request_tree.yview)
        request_tree.configure(yscrollcommand=scrollbar.set)
        
        request_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        def refresh_list(search_term=""):
            try:
                conn = sqlite3.connect(REQUEST_SHEET_DB)
                cursor = conn.cursor()
                
                if search_term:
                    cursor.execute("""
                        SELECT request_sheet_no, issue_date, customer_name, po_number, status
                        FROM request_sheets
                        WHERE request_sheet_no LIKE ? OR customer_name LIKE ? OR po_number LIKE ?
                        ORDER BY created_date DESC
                    """, (f"%{search_term}%", f"%{search_term}%", f"%{search_term}%"))
                else:
                    cursor.execute("""
                        SELECT request_sheet_no, issue_date, customer_name, po_number, status
                        FROM request_sheets
                        ORDER BY created_date DESC
                    """)
                
                rows = cursor.fetchall()
                conn.close()
                
                # Clear tree
                for item in request_tree.get_children():
                    request_tree.delete(item)
                
                # Populate tree
                for row in rows:
                    request_tree.insert("", "end", values=row)
                    
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load request sheets: {e}")
        
        def load_selected():
            selection = request_tree.selection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a request sheet to load.")
                return
            
            request_no = request_tree.item(selection[0])['values'][0]
            popup.destroy()
            self.load_request_sheet_by_number(request_no)
        
        request_tree.bind('<Double-Button-1>', lambda e: load_selected())
        
        # Buttons
        button_frame = tk.Frame(popup, bg='white')
        button_frame.pack(fill='x', padx=15, pady=(0, 15))
        
        tk.Button(button_frame, text="📂 Load Selected", command=load_selected, bg="#16a34a", fg="white",
                 font=("Segoe UI", 11, "bold"), padx=20, pady=5, relief='raised', borderwidth=2).pack(side='left', padx=5)
        
        tk.Button(button_frame, text="❌ Close", command=popup.destroy, bg="#6b7280", fg="white",
                 font=("Segoe UI", 11, "bold"), padx=20, pady=5, relief='raised', borderwidth=2).pack(side='right', padx=5)
        
        # Initial load
        refresh_list()
    
    def add_material_row(self):
        """Add a new material row to the table"""
        row_num = self.next_row_number
        
        # Create row widgets
        row_data = {
            'row_num': row_num,
            'widgets': {}
        }
        
        # Row number label
        tk.Label(self.material_frame, text=str(row_num), bg='white', font=("Segoe UI", 10),
                relief='solid', borderwidth=1, width=5).grid(row=row_num, column=0, sticky='ew', padx=1, pady=1)
        
        # Model No entry
        model_entry = tk.Entry(self.material_frame, width=25, font=("Segoe UI", 10))
        model_entry.grid(row=row_num, column=1, sticky='ew', padx=1, pady=1)
        model_entry.bind('<KeyRelease>', lambda e: self.on_model_no_change(e, row_data))
        row_data['widgets']['model_no'] = model_entry
        
        # Barcode entry (auto-filled from model number)
        barcode_entry = tk.Entry(self.material_frame, width=30, font=("Segoe UI", 10), state='readonly')
        barcode_entry.grid(row=row_num, column=2, sticky='ew', padx=1, pady=1)
        row_data['widgets']['barcode'] = barcode_entry
        
        # Qty/Weight entry
        qty_entry = tk.Entry(self.material_frame, width=15, font=("Segoe UI", 10))
        qty_entry.grid(row=row_num, column=3, sticky='ew', padx=1, pady=1)
        row_data['widgets']['qty_weight'] = qty_entry
        
        # Delete button
        delete_btn = tk.Button(self.material_frame, text="Remove", bg="#dc2626", fg="white",
                              font=("Segoe UI", 9), command=lambda: self.remove_material_row(row_num))
        delete_btn.grid(row=row_num, column=4, padx=2, pady=2)
        row_data['widgets']['delete_btn'] = delete_btn
        
        self.material_rows.append(row_data)
        self.next_row_number += 1
    
    def remove_material_row(self, row_num):
        """Remove a material row from the table"""
        for row_data in self.material_rows:
            if row_data['row_num'] == row_num:
                # Destroy all widgets in the row
                for widget in row_data['widgets'].values():
                    widget.destroy()
                # Remove from list
                self.material_rows.remove(row_data)
                break
    
    def save_request_sheet(self):
        """Save request sheet to database"""
        # Validate inputs
        request_no = self.request_no_entry.get().strip()
        issue_date = self.issue_date_entry.get().strip()
        customer_name = self.customer_name_entry.get().strip()
        qty_sensor = self.qty_sensor_entry.get().strip()
        po_number = self.po_number_entry.get().strip()
        
        if not all([request_no, issue_date, customer_name, qty_sensor, po_number]):
            messagebox.showwarning("Validation Error", "Please fill in all required fields: Request No, Issue Date, Customer Name, Qty of Sensor, and PO Number.")
            return
        
        try:
            qty_sensor = int(qty_sensor)
        except ValueError:
            messagebox.showwarning("Validation Error", "Qty of Sensor must be a number.")
            return
        
        if not self.material_rows:
            messagebox.showwarning("Validation Error", "Please add at least one material item.")
            return
        
        # Collect material items
        material_items = []
        for row_data in self.material_rows:
            model_no = row_data['widgets']['model_no'].get().strip()
            barcode = row_data['widgets']['barcode'].get().strip()
            qty_weight = row_data['widgets']['qty_weight'].get().strip()
            
            if model_no and qty_weight:
                material_items.append({
                    'item_no': row_data['row_num'],
                    'model_no': model_no,
                    'barcode': barcode,
                    'qty_weight': qty_weight
                })
        
        if not material_items:
            messagebox.showwarning("Validation Error", "Please fill in at least one complete material item (Model No and Qty/Weight).")
            return
        
        try:
            conn = sqlite3.connect(REQUEST_SHEET_DB)
            cursor = conn.cursor()
            
            # Check if request sheet exists (update vs insert)
            cursor.execute("SELECT id FROM request_sheets WHERE request_sheet_no=?", (request_no,))
            existing = cursor.fetchone()
            
            if existing:
                # Update existing
                cursor.execute("""
                    UPDATE request_sheets 
                    SET issue_date=?, customer_name=?, qty_of_sensor=?, po_number=?, item_code=?,
                        checker_name=?, checker_date=?, approver_name=?, approver_date=?
                    WHERE request_sheet_no=?
                """, (
                    issue_date, customer_name, qty_sensor, po_number, 
                    self.item_code_entry.get().strip(),
                    self.checker_name_entry.get().strip(), self.checker_date_entry.get().strip(),
                    self.approver_name_entry.get().strip(), self.approver_date_entry.get().strip(),
                    request_no
                ))
                
                # Delete old material items
                cursor.execute("DELETE FROM material_items WHERE request_sheet_no=?", (request_no,))
            else:
                # Insert new
                cursor.execute("""
                    INSERT INTO request_sheets 
                    (request_sheet_no, issue_date, customer_name, qty_of_sensor, po_number, item_code,
                     checker_name, checker_date, approver_name, approver_date, created_by, created_date, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    request_no, issue_date, customer_name, qty_sensor, po_number,
                    self.item_code_entry.get().strip(),
                    self.checker_name_entry.get().strip(), self.checker_date_entry.get().strip(),
                    self.approver_name_entry.get().strip(), self.approver_date_entry.get().strip(),
                    "Admin", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Pending"
                ))
            
            # Insert material items
            for item in material_items:
                cursor.execute("""
                    INSERT INTO material_items (request_sheet_no, item_no, model_no, barcode, qty_weight)
                    VALUES (?, ?, ?, ?, ?)
                """, (request_no, item['item_no'], item['model_no'], item['barcode'], item['qty_weight']))
            
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Success", f"Request sheet {request_no} saved successfully!")
            
        except sqlite3.IntegrityError as e:
            messagebox.showerror("Database Error", f"Request sheet number already exists: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save request sheet: {e}")
    
    def load_request_sheet_by_number(self, request_no):
        """Load a request sheet by its number"""
        try:
            conn = sqlite3.connect(REQUEST_SHEET_DB)
            cursor = conn.cursor()
            
            # Load master data
            cursor.execute("SELECT * FROM request_sheets WHERE request_sheet_no=?", (request_no,))
            sheet_data = cursor.fetchone()
            
            if not sheet_data:
                messagebox.showwarning("Not Found", "Request sheet not found.")
                return
            
            # Load material items
            cursor.execute("SELECT * FROM material_items WHERE request_sheet_no=? ORDER BY item_no", (request_no,))
            material_data = cursor.fetchall()
            conn.close()
            
            # Clear form
            self.clear_form()
            
            # Fill header fields
            self.request_no_entry.insert(0, sheet_data[1])
            self.issue_date_entry.delete(0, tk.END)
            self.issue_date_entry.insert(0, sheet_data[2])
            self.customer_name_entry.insert(0, sheet_data[3])
            self.qty_sensor_entry.insert(0, str(sheet_data[4]))
            self.po_number_entry.insert(0, sheet_data[5])
            self.item_code_entry.insert(0, sheet_data[6] if sheet_data[6] else "")
            self.checker_name_entry.insert(0, sheet_data[7] if sheet_data[7] else "")
            self.checker_date_entry.insert(0, sheet_data[8] if sheet_data[8] else "")
            self.approver_name_entry.insert(0, sheet_data[9] if sheet_data[9] else "")
            self.approver_date_entry.insert(0, sheet_data[10] if sheet_data[10] else "")
            
            # Fill material items
            for item in material_data:
                self.add_material_row()
                # Find the last added row
                last_row = self.material_rows[-1]
                last_row['widgets']['model_no'].insert(0, item[3])
                # Temporarily enable barcode entry to insert value
                last_row['widgets']['barcode'].config(state='normal')
                last_row['widgets']['barcode'].insert(0, item[4] if item[4] else "")
                last_row['widgets']['barcode'].config(state='readonly')
                last_row['widgets']['qty_weight'].insert(0, item[5])
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load request sheet: {e}")
    
    def new_request_sheet(self):
        """Create a new request sheet"""
        self.clear_form()
        self.generate_request_number()
        self.issue_date_entry.delete(0, tk.END)
        self.issue_date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
    
    def clear_form(self):
        """Clear all form fields"""
        self.request_no_entry.delete(0, tk.END)
        self.issue_date_entry.delete(0, tk.END)
        self.customer_name_entry.delete(0, tk.END)
        self.qty_sensor_entry.delete(0, tk.END)
        self.po_number_entry.delete(0, tk.END)
        self.item_code_entry.delete(0, tk.END)
        self.checker_name_entry.delete(0, tk.END)
        self.checker_date_entry.delete(0, tk.END)
        self.approver_name_entry.delete(0, tk.END)
        self.approver_date_entry.delete(0, tk.END)
        
        # Clear material rows
        for row_data in self.material_rows:
            for widget in row_data['widgets'].values():
                widget.destroy()
        self.material_rows = []
        self.next_row_number = 1
    
    def delete_request_sheet(self):
        """Delete the currently loaded request sheet"""
        request_no = self.request_no_entry.get().strip()
        
        if not request_no:
            messagebox.showwarning("No Selection", "Please load a request sheet to delete.")
            return
        
        confirm = messagebox.askyesno("Confirm Delete", 
                                     f"Are you sure you want to delete request sheet {request_no}?\n\nThis action cannot be undone.")
        
        if not confirm:
            return
        
        try:
            conn = sqlite3.connect(REQUEST_SHEET_DB)
            cursor = conn.cursor()
            
            cursor.execute("DELETE FROM material_items WHERE request_sheet_no=?", (request_no,))
            cursor.execute("DELETE FROM request_sheets WHERE request_sheet_no=?", (request_no,))
            
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Success", f"Request sheet {request_no} deleted successfully.")
            self.clear_form()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete request sheet: {e}")
    
    def export_to_excel(self):
        """Export request sheet to Excel with formatted layout"""
        request_no = self.request_no_entry.get().strip()
        
        if not request_no:
            messagebox.showwarning("No Data", "Please create or load a request sheet to export.")
            return
        
        # Ask user for save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{request_no}.xlsx",
            title="Save Request Sheet"
        )
        
        if not file_path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Request Sheet"
            
            # Styling
            title_font = Font(size=16, bold=True)
            header_font = Font(size=11, bold=True)
            normal_font = Font(size=10)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
            blue_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            
            # Title
            ws.merge_cells('A1:E1')
            ws['A1'] = "Request Sheet"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Header section - Customer Info
            ws.merge_cells('A3:C3')
            ws['A3'] = "Customer Information"
            ws['A3'].font = header_font
            
            ws['A4'] = "Customer Name:"
            ws['B4'] = self.customer_name_entry.get()
            ws['A5'] = "Qty of Sensor:"
            ws['B5'] = self.qty_sensor_entry.get()
            ws['A6'] = "PO Number:"
            ws['B6'] = self.po_number_entry.get()
            ws['A7'] = "Item Code:"
            ws['B7'] = self.item_code_entry.get()
            
            # Header section - Request Info (right side)
            ws['D3'] = "Issue Date:"
            ws['E3'] = self.issue_date_entry.get()
            ws['D4'] = "Request Sheet No:"
            ws['E4'] = request_no
            
            # Material List section
            ws.merge_cells('A9:E9')
            ws['A9'] = "Material List"
            ws['A9'].font = header_font
            
            # Material table headers
            headers = ['No', 'Model No', 'Barcode (Model No)', 'Qty/Weight']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=10, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = blue_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Material items data
            row_num = 11
            for idx, row_data in enumerate(self.material_rows, start=1):
                model_no = row_data['widgets']['model_no'].get()
                barcode = row_data['widgets']['barcode'].get()
                qty_weight = row_data['widgets']['qty_weight'].get()
                
                if model_no and qty_weight:
                    ws.cell(row=row_num, column=1, value=idx).border = thin_border
                    ws.cell(row=row_num, column=2, value=model_no).border = thin_border
                    ws.cell(row=row_num, column=3, value=barcode).border = thin_border
                    ws.cell(row=row_num, column=4, value=qty_weight).border = thin_border
                    row_num += 1
            
            # Approval section
            approval_row = row_num + 2
            
            # Checker
            ws.cell(row=approval_row, column=1, value="Checker").font = header_font
            ws.cell(row=approval_row + 1, column=1, value="Name:")
            ws.cell(row=approval_row + 1, column=2, value=self.checker_name_entry.get())
            ws.cell(row=approval_row + 2, column=1, value="Date:")
            ws.cell(row=approval_row + 2, column=2, value=self.checker_date_entry.get())
            
            # Approver
            ws.cell(row=approval_row, column=4, value="Approver").font = header_font
            ws.cell(row=approval_row + 1, column=4, value="Name:")
            ws.cell(row=approval_row + 1, column=5, value=self.approver_name_entry.get())
            ws.cell(row=approval_row + 2, column=4, value="Date:")
            ws.cell(row=approval_row + 2, column=5, value=self.approver_date_entry.get())
            
            # Footer
            footer_row = approval_row + 4
            ws.merge_cells(f'A{footer_row}:E{footer_row}')
            ws[f'A{footer_row}'] = "INTERNAL USE ONLY"
            ws[f'A{footer_row}'].alignment = Alignment(horizontal='center')
            ws[f'A{footer_row}'].font = Font(size=9, italic=True)
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            
            # Save workbook
            wb.save(file_path)
            messagebox.showinfo("Export Success", f"Request sheet exported successfully to:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export request sheet: {e}")


def main():
    """Main entry point"""
    root = tk.Tk()
    app = MaterialRequestSheetSystem(root)
    root.mainloop()

if __name__ == "__main__":
    main()
